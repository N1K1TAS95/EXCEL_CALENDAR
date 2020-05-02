from openpyxl import load_workbook

wb = load_workbook('ORE NIKITA.xlsx')

ws = wb.create_sheet('PROVA')

wb.save('ORE NIKITA.xlsx')
