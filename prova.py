from datetime import date
from calendar import monthrange
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

week_days_names = {
    0: 'Lunedì',
    1: 'Martedì',
    2: 'Mercoledì',
    3: 'Giovedì',
    4: 'Venerdì',
    5: 'Sabato',
    6: 'Domenica'
}

year = int(input("Inserisci l'anno > "))
month = int(input('Inserisci il mese > '))

days = [[date(year, month, x).strftime("%d/%m/%Y"), week_days_names[date(year, month, x).weekday()]]
        for x in range(1, monthrange(year, month)[1] + 1)]

print(days)

wb = load_workbook('ORE NIKITA.xlsx')
ws = wb.create_sheet('PROVA')

for x in range(1, len(days) + 1):
    ws.cell(1, x, days[x - 1][1]).border = thin_border
    ws.cell(2, x, days[x - 1][0]).border = thin_border

wb.save('ORE NIKITA.xlsx')
    
